import pandas as pd
import yfinance as yf
import xlsxwriter
import openpyxl

from openpyxl import load_workbook

yf.pdr_override()

class ticker: 
    """ a base class represnting a ticker and associated financial metadata""" 

    def __init__(self, symbol):
        self.symbol = symbol
        self.marketCap = 0

    def populateTicker(self):
        # consolidate yfinance calls 
        
        print ('starting INFO lookup...')
        self.info = yf.Ticker(self.symbol).info
        print ('DONE!')

        print ('starting FINANCIALS lookup...')
        self.financials = yf.Ticker(self.symbol).financials
        print ('DONE!')
        
        print ('starting BALANCE SHEET lookup...')
        self.balanceSheet = yf.Ticker(self.symbol).balance_sheet
        print ('DONE!')

        print ('getting 1d quote...')
        self.history1d = yf.Ticker(self.symbol).history(period="1d")
        print ('DONE!')

        self.populateKeyStats()
        self.populateCompanyInformation()
        self.populateQuote()
        self.populateAnnualFinancials()

        # TODO implement below calls with actual data
        self.companyName = 1000
        self.exchange = 1000
        self.salesPerShare = 1111
        self.cashflowPerShare = 1111
        self.insiderOwnership = 1111
        self.insiderBuysTTM = 1111
        self.stockBuyBack = 1111
        self.EPSRank = 1111
        self.RPSRank = 1111
        self.earningsPerShare5y = 1111
        self.revenue5y = 1111
        self.latestPrice5y = 1111
        self.projectedSales = 1111
        self.projectedHigh = 1111
        self.projectedLow = 1111
        self.starsFairVal = 1111
        self.currentPE = 1111
        self.averagePE = 1111
        self.priceToSales = 1111
        self.currentRatio = 1111
        self.quickRatio = 1111
        self.returnOnEquity = 1000        
        self.priceToBook = 1000
        self.sharesOutstanding = 1000
        
        # Dependencies -> populateAnnualFinancials() &&  populateKeyStats
        self.earningsPerShare = self.netIncome / self.sharesOutstanding
        
        print ('Printing to Excel...')
        self.printExcel()
        print ('DONE! ')

        #this P/E is invalid since the EPS is from fiscal 2017 and the price is the latest market price! 
        #self.priceToEarnings = self.latestPrice / self.earningsPerShare
            
    def populateKeyStats(self):
        """ populates the ticker with: 
        Market Capitalization, 
        dividend yield, 
        TTM Return on Equity, 
        price/book """
        
       # pd = dataFeed.returnKeyStats(self.symbol)
        self.marketCap = self.info['marketCap']
        self.dividendYield = self.info['dividendYield']

    def populateCompanyInformation(self): 
        """ populate the ticker with:
        Name 
        Exchange
        Industry
        Description
        Website """
        #pd = dataFeed.returnCompanyInformation(self.symbol)

        
        self.sector = self.info['sector']
        self.description = self.info['longBusinessSummary']
        self.website = self.info['website']
        self.week52High = self.info['fiftyTwoWeekHigh']
        self.week52Low = self.info['fiftyTwoWeekLow']
        self.latestVolume = self.info['averageVolume']
        self.netProfitMargin = self.info['profitMargins']

    def populateQuote(self):
        """ populate ticker with realtime quote info """
        #pd = dataFeed.returnQuote(self.symbol)
        self.latestPrice = self.history1d['Close'][0]


    def populateAnnualFinancials(self):
        """ populate ticker with annual financials """
        #pd = dataFeed.returnFinancials_Annual(self.symbol)
        self.revenue = self.financials[self.financials.columns[0]]['Total Revenue']
        self.netIncome = self.financials[self.financials.columns[0]]['Net Income']
        
        self.cash = self.balanceSheet[self.balanceSheet.columns[0]]['Cash']
        self.currentLiabilities = self.balanceSheet[self.balanceSheet.columns[0]]['Total Current Liabilities']
        self.shareholderEquity = self.balanceSheet[self.balanceSheet.columns[0]]['Total Stockholder Equity']
        self.commonStock = self.balanceSheet[self.balanceSheet.columns[0]]['Common Stock']

    # Convert ticker object into a dataframe for simplified data manipulation 
    # TODO is this the simplest way? 
    def tickerToDataframe(self):

        
        return pd.DataFrame({'Data' : [self.symbol, self.sector, self.latestPrice, self.week52High, self.week52Low, self.marketCap, self.latestVolume, self.commonStock, self.revenue, self.shareholderEquity, self.netIncome, self.netProfitMargin,self.cash, self.currentLiabilities, self.salesPerShare, self.cashflowPerShare, self.earningsPerShare, self.dividendYield, self.returnOnEquity, self.insiderOwnership, self.insiderBuysTTM, self.stockBuyBack, self.EPSRank, self.RPSRank, self.earningsPerShare5y, self.revenue5y, self.latestPrice5y, self.projectedSales, self.projectedHigh, self.projectedLow, self.starsFairVal, self.currentPE, self.averagePE, self.priceToSales, self.priceToBook, self.currentRatio, self.quickRatio]})
    
    def printExcel(self):
       # self.tickerToDataframe()
       # covert the tickercalss into a dataframe to simplify writing to Excel 
        testdf = self.tickerToDataframe()

        # load up the file 
        # TODO catch file not found error
        # TODO filename from config 
        myFilename = 'testFile.xlsx'
        mySheetName = 'Sheet1'
        myExcelFile = load_workbook(filename = myFilename)
        myWorksheet = myExcelFile['Sheet1']

        nextRow = 1
        nextCol = myWorksheet.max_column + 1

        writer = pd.ExcelWriter(myFilename, engine = 'openpyxl')
        writer.book = myExcelFile
        writer.sheets = {ws.title: ws for ws in myExcelFile.worksheets}

        #writer.book = load_workbook('testFile.xlsx')

        
        

        testdf.to_excel(writer, sheet_name=mySheetName, startrow = nextRow, startcol = nextCol, header=False, index=False, )
        #print ( writer.sheets['Sheet1'].max_row )

        writer.save()
        
        #...................SAMPLE CODE FOR XLSXWRITER library        
        
        #wb = xlsxwriter.Workbook('testFile.xlsx')
        #ws = wb.add_worksheet()
        #ws.write(0,0, 'Hello from Python!', startrow=5)
        #wb.close()
        # It is also possible to write the dataframe without the header and index.
        #df4.to_excel(writer, sheet_name='Sheet1', startrow=7, startcol=4, header=False, index=False)

